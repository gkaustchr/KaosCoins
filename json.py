from openpyxl import load_workbook
import json

def xlsx_to_json(filepath):
    """
    Converte um arquivo XLSX para um dicionário JSON.

    Args:
        filepath: Caminho para o arquivo XLSX.

    Returns:
        Um dicionário Python representando o JSON, ou None se ocorrer um erro.
    """
    try:
        workbook = load_workbook(filepath, data_only=True) # data_only=True para ler valores e não fórmulas
        sheet = workbook.active  # Seleciona a primeira planilha

        dt = {"catalogo": []}
        keys = [cell.value for cell in sheet[1] if cell.value is not None] # Pega a linha 1 (índice 0)
        data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True): # Começa da segunda linha (assumindo a primeira é o cabeçalho)
            for i, value in enumerate(row):
                # data[keys[i]] = value
                # continue
                if( ";" in str(value)):
                    if(":" in value and not "https:" in value):
                        ficha_data = value.replace('"', '') 
                        # Divide a string em pares chave-valor
                        pares_chave_valor = ficha_data.split(';')
                        # Cria um dicionário
                        ficha_dict = {}
                        for par in pares_chave_valor:
                            chave, valor = par.split(':', 1)  # split(': ', 1) limita a 1 divisão
                            chave = chave.replace(" ", "")
                            ficha_dict[chave] = valor.strip().replace('"', "'").replace('None', '""') #
                        data[keys[i]] = ficha_dict
                    else:
                        data[keys[i]] = value.strip().split(";")
                else:
                    if value == None:
                        data[keys[i]] = ""
                    else:
                        data[keys[i]] = value

            dt["catalogo"].append(data)
            data = {}

        return dt
    except FileNotFoundError:
        print(f"Erro: Arquivo não encontrado em '{filepath}'")
        return None
    except Exception as e:
        print(f"Erro ao processar o arquivo: {e}")
        return None


# Exemplo de uso:
filepath = "catalogo.xlsx"  # Substitua pelo caminho do seu arquivo
json_data = xlsx_to_json(filepath)

if json_data:
    print(str(json_data).replace("'", '"').replace("False", "false").replace("True", "true")) # Imprime o JSON formatado
    # print(json.dumps(json_data, indent=4)) # Imprime o JSON formatado